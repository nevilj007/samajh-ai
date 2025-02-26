from fastapi import FastAPI, Form, Request
from ollama import chat
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from fastapi.responses import HTMLResponse
from typing import Dict, Any, List
from database import get_db_connection
import requests
import fitz
import os
from dotenv import load_dotenv
import tempfile
from ollama import Client


app = FastAPI()
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")
# Connect to the Ollama container inside Docker
ollama_client = Client(host='http://localhost:11434')


# In-memory storage for conversations
conversations: Dict[str, Dict[str, Any]] = {}
load_dotenv()
api_key = os.getenv("BLAND_API_KEY")


def download_pdf(pdf_url, save_path="downloaded.pdf"):
    """
    Downloads a PDF file from a given URL.
    :param pdf_url: URL of the PDF
    :param save_path: Local path to save the downloaded PDF
    :return: Path to the downloaded PDF
    """
    response = requests.get(pdf_url, stream=True)
    if response.status_code == 200:
        with open(save_path, "wb") as pdf_file:
            pdf_file.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download PDF. Status code: {response.status_code}")

def extract_text_blocks(pdf_path):
    """
    Extracts text from a PDF file while maintaining order using PyMuPDF (blocks method).
    :param pdf_path: Path to the PDF file
    :return: Extracted and ordered text
    """
    doc = fitz.open(pdf_path)
    text = ""

    for page in doc:
        blocks = page.get_text("blocks")  # Extract text as blocks
        blocks.sort(key=lambda b: (b[1], b[0]))  # Sort by (y, x) coordinates

        for block in blocks:
            text += block[4] + "\n"

    return text.strip()





# Initialize the agent for prompt generation
db_url = "postgresql+psycopg://ai:ai@localhost:5532/ai"
prompt = {}
class Transcript(BaseModel):
    id: int
    created_at: str
    text: str
    user: str
def transcript_to_dict(transcript):
    return {
        'id': transcript.id,
        'created_at': transcript.created_at,
        'text': transcript.text,
        'user': transcript.user
    }

class WebhookData(BaseModel):
    call_id: str
    transcripts: List[Transcript]
    concatenated_transcript: str
    to:str
    summary: str
    call_length: float
    price: float


class Message(BaseModel):
    message: str




# Add new Pydantic model for prompt generation
class PromptData(BaseModel):
    call_output: str = Form(...)
    suggestions: str = Form(...)
    modified_prompt: str = Form(None)
@app.get("/")
async def home(request: Request):
    return templates.TemplateResponse("prompt_generator2.html", {"request": request})

# Update your prompt generator endpoint
@app.post("/generate_prompt")
async def generate_prompt(
        request: Request,
        knowledge_base_url: str = Form(...),
        phone_number: str = Form(...),
        questions: str = Form(...),
        call_output: str = Form(...),
        suggestions: str = Form(...)
):

    global question_lists
    question_list=questions.split('\n')
    question_lists=question_list
    generation_input = f"""
        questions:{question_list}
        Call Output: {call_output}
        Suggestions: {suggestions}
        ###instructions##
        I have a caller agent who calls the customers and asks the questions mentioned in {question_list} to them understands their preferences and then suggest the recommendations available in their knwoledge base.
        Your role is to  generate a comprehensive prompt in 250 words or more for the caller agent based on the below information.
        Analyse  {suggestions} ,{call_output},{question_list} for generating the prompt.
        Always include {call_output}  in your prompt as the motive of calling the customer.
        Always stricty the content of the prompt should never have any product specifications  not mentioned in  {suggestions} ,{call_output},{question_list}.
        ALWAYS include these lines in the prompt:'be more creative,dont use long and boring sentences,and dont repeat same sentences.also dont hallucinate'.
        Also mention the communication style of the caller agent with the customer by deeply analysing the  {suggestions} ,{call_output},{question_list}.
        Remember the {question_list} contains things you dont know so dont assume that you know of the details regarding the customer mentioned in {question_list}.
        Strictly generate the prompt content . DONT INCLUDE extra text or context like  'Here is your prompt' or 'Based on the provided context, I will create a detailed prompt for the caller agent to interact with customers'.
        Please generate a comprehensive prompt in 150 words or more for the caller agent based on this information.
        """

    def prompt_generator(user_message: str, caller_id: str):
        if caller_id not in prompt:
            try:
                response = ollama_client.chat(
                    model="llama3.1:latest",
                    messages=[{'role': 'user', 'content': user_message}]
                )
                if response.get('message'):
                    return response['message']['content']
                else:
                    return "I did not receive any valid answer from the convo."

            except Exception as e:
                print(f"Error communicating with Ollama: {e}")
                return "Error generating a response."

    generated_prompt = prompt_generator(generation_input, "1")
    return templates.TemplateResponse("prompt_generator2.html", {
         "request": request,
        "generated_prompt": generated_prompt,
        "phone_number": phone_number,
        "knowledge_base_url": knowledge_base_url,
        "questions": questions,
        "call_output": call_output,
        "suggestions": suggestions
    })


# Update your make_call endpoint
@app.post("/make_call")
async def make_call(
        request: Request,
        phone_number: str = Form(...),
        knowledge_base_url: str = Form(...),
        final_prompt: str = Form(...),
        questions: str = Form(...)
):
    # 🔹 Step 1: Download the PDF
    pdf_path = download_pdf(knowledge_base_url)

    # 🔹 Step 2: Extract text using PyMuPDF (default method)
    formatted_text = extract_text_blocks(pdf_path)

    # 🔹 Delete the downloaded PDF after processing (Optional)
    os.remove(pdf_path)
    headers = {
        'authorization': api_key
    }

    question_list = questions.split('\n')
    prompter = f"""
    You are Jake for the service provider mentioned in the {knowledge_base_url}.Give a brief intro before asking the question for him/her to be familiar with you. Your knowledge is strictly limited to the information provided in the knowledge base.

    Instructions:
    1.The knowledge base is {formatted_text}. 
    2. If asked about any topic not covered in the knowledge base, respond with: "I'm sorry, but I don't have information about that. I can only provide details about our products and services as mentioned in our knowledge base."
    3. Only answer questions about products, services, and information mentioned in the knowledge base.
    4. Follow the {final_prompt} format exactly.
    5. Do not make up or infer any information not explicitly stated in the knowledge base.

    Remember: If you're unsure or the information isn't in the knowledge base, always say you don't have that information.
    """

    data = {
        'phone_number': phone_number,
        'task': f"Always greet the caller .STRICTLY Ask ALL the questions in the {question_list} one at a time, waiting for a response before moving to the next question.DO NOT LEAVE ANY QUESTION FROM {question_list}UNASKED.also when you get interrupted by user when you ask a question address his query and aks the question after it .ENSURE TO ASK ALL THE QUESTION TILL YOU GET A VALID REPLY .NO IS ALSO A VALID REPLY.Only use information from this PDF. If the information isn't in the knowledge base, say that you don't have that information. Strictly follow the instructions in the {prompter}.Always end the call only if customer wishes to end it",

        'prompt': prompter,
        'webhook': 'https://c1dd-103-37-201-225.ngrok-free.app/webhook',

        'record': True,
        'reduce_latency': True,
        'amd': True,
        'model': 'base'  # Using the 'base' model as it follows scripts/procedures most effectively
    }

    response = requests.post('https://api.bland.ai/v1/calls', json=data, headers=headers)
    if response.status_code != 200:
        print(f"Error: {response.status_code}")
        print(response.text)
    else:
        print(response.json())
    return {"message": "Call initiated", "call_id": response.json().get('call_id'), "phone number": phone_number,
            "question": question_list,"webhook":data['webhook']}


import json


def json_to_text(text_file_path):
    # Load the JSON data from the file as a string
    with open(text_file_path, 'r', encoding='utf-8') as file:
        data_str = file.read()

    # Parse the JSON string into a dictionary
    data = json.loads(data_str)

    # Prepare the text content by iterating over the dictionary items
    text_content = ""
    for key, value in data.items():
        text_content += f"{key}: {value}\n"

    # Write the text content to an output file (or overwrite the original file if desired)

    with open(text_file_path, 'w', encoding='utf-8') as file:
        file.write(text_content)

    print(f"Converted JSON to text and saved to {text_file_path}")
    return text_file_path





# Function to read text from a file
def read_text_file(filepath):
    with open(filepath, 'rb') as file:
        content = file.read()

    try:
        return content.decode('utf-8')
    except UnicodeDecodeError:
        # Replace problematic characters with a replacement character
        return content.decode('utf-8', errors='replace')


# Main function to interact with Gemini
def query_gemini(file_path):
    file_text = read_text_file(file_path)

    if not file_text:
        print("No content provided or Error reading file.")
        return

    prompt_1 = f"""
    ###Instructions###
    Assistant calls the user and has a conversation with him.   
    Below is the conversation between the user and assistant:
    {file_text}

    Based on this conversation, generate a detailed summary extracting all important points.
    Also, mention whether the user has answered all the questions asked by the assistant. If not, specifically list the unanswered questions.
    """

    try:
        response = ollama_client.chat(
            model="llama3.1:latest",
            messages=[{'role': 'user', 'content': prompt_1}]
        )
        if response.get('message'):
            return response['message']['content']
        else:
            return "I did not receive any valid answer from the conversation."

    except Exception as e:
        print(f"Error communicating with Ollama: {e}")
        return "Error processing conversation summary."



def final_reply(reply_1, question):
    prompt_2 = f"""
    ###Instructions###

    You will be given a detailed summary of the conversation between a user and an assistant, along with a specific question asked by the assistant. Your task is to extract and output the user's response in the most concise form possible—ideally a single word or phrase.

    Guidelines:  
    - Go through the entire conversation and extract the user's exact responses.
    - Ensure all aspects of the user's answer are included, even if mentioned separately.
    - Convert the response into the most concise form possible (a single word or phrase).
    - Correct any grammatical mistakes in the user response by analyzing the assistant's reply.
    - Do not include symbols like "*".
    - Do not include introductory phrases like "The user's response was...".
    - If the user has not responded to the question, output 'NOT AVAILABLE'.
    - Include specific mentions of any product if relevant to the question.
    - Do NOT include anything else.
    - Base your answer strictly on the provided conversation—do not infer or hallucinate responses.

    Here is a detailed summary of the conversation:
    {reply_1}

    And here is the question:
    {question}
    """

    try:
        response = ollama_client.chat(
            model="llama3.1:latest",
            messages=[{'role': 'user', 'content': prompt_2}]
        )
        if response.get('message'):
            return response['message']['content']
        else:
            return "I did not receive any valid answer."

    except Exception as e:
        print(f"Error communicating with Ollama: {e}")
        return "Error processing final reply."




# 🔹 Save conversation data to PostgreSQL
async def save_conversation(call_id, transcripts, summary, audio_url,phone_number):
    conn = await get_db_connection()
    await conn.execute(
        """
        INSERT INTO conversations (call_id, transcripts, summary, audio_url,phone_number)
        VALUES ($1, $2, $3, $4,$5)
        ON CONFLICT (call_id) DO UPDATE
        SET transcripts = $2, summary = $3, audio_url = $4,phone_number=$5;
        """,
        call_id, transcripts, summary, audio_url,phone_number
    )
    await conn.close()

# 🔹 Save each question-response pair to PostgreSQL
async def save_call_responses(call_id, question, response):
    conn = await get_db_connection()
    await conn.execute(
        """
        INSERT INTO call_responses (call_id, question, response)
        VALUES ($1, $2, $3);
        """,
        call_id, question, response
    )
    await conn.close()
@app.post("/webhook")
async def webhook(data: WebhookData):
    call_id = data.call_id
    if call_id:
        conversations[call_id] = {
            'transcripts': [transcript_to_dict(t) for t in data.transcripts],
            'summary': data.summary,
            'to_phone': data.to



        }
        print(conversations[call_id])
        b=conversations[call_id]['to_phone']
        print(b)
        # Extract the conversation from the transcripts
        conversation = []
        for transcript in conversations[call_id]['transcripts']:
            conversation.append(f"{transcript['user'].capitalize()}: {transcript['text']}")

        # Join the conversation lines
        conversation_text = "\n".join(conversation)


        # 🔥 Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".txt", mode="w", encoding="utf-8") as temp_file:
            temp_file.write(conversation_text)
            temp_file_path = temp_file.name  # Get file path before closing
        BLAND_API_KEY = api_key



        url = f"https://api.bland.ai/v1/calls/{call_id}/recording"

        headers = {
            "authorization": BLAND_API_KEY,
            "accept": "application/json"
        }

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            if data["status"] == "success":
                print(f"Audio URL: {data['url']}")
            else:
                print(f"Error: {data['message']}")
        else:
            print(f"Error: Unable to retrieve audio URL. Status code: {response.status_code}")
        transcripts=read_text_file(temp_file_path)
        summary=query_gemini(temp_file_path)
        audio_url=data['url']
        # Save to PostgreSQL
        await save_conversation(call_id,transcripts,summary,audio_url,b)

        print(summary)




        print(len(question_lists))
        print(question_lists)
        for i in range(len(question_lists)):
            if question_lists[i].strip():  #  Check if question is NOT empty
                response = final_reply(summary, question_lists[i])
                print(question_lists[i])
                print(response)
                await save_call_responses(call_id, question_lists[i], response)


        os.remove(temp_file_path)
        return {"message": f"Data saved to PostgreSQL"}



# New route to list Excel files
@app.get("/files", response_class=HTMLResponse)
async def list_files(request: Request):
    conn = await get_db_connection()

    # 🔹 Get all available call IDs from PostgreSQL
    call_ids = await conn.fetch("SELECT call_id FROM conversations")

    await conn.close()

    # 🔹 Extract call_id values
    call_list = [record["call_id"] for record in call_ids]

    return templates.TemplateResponse(
        "file_list.html", {"request": request, "files": call_list}
    )


# New route to download files
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook

@app.get("/download/{call_id}")
async def download_file(call_id: str):
    conn = await get_db_connection()

    # 🔹 Fetch Data from PostgreSQL
    responses = await conn.fetch(
        "SELECT question, response FROM call_responses WHERE call_id = $1", call_id
    )
    summary = await conn.fetchval(
        "SELECT summary FROM conversations WHERE call_id = $1", call_id
    )

    audio_url = await conn.fetchval(
        "SELECT audio_url FROM conversations WHERE call_id = $1", call_id
    )

    phone_number=await conn.fetchval(
        "SELECT phone_number FROM conversations WHERE call_id=$1",call_id
    )


    await conn.close()

    # 🔹 Generate Excel File in Memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Data"

    # 🔹 Add Questions & Responses
    for i, row in enumerate(responses):
        ws[f'A{i+1}'] = row['question']
        ws[f'B{i+1}'] = row['response']

    # 🔹 Add Summary, Phone Number, and Audio URL
    ws.append([])
    ws.append(["Summary", summary if summary else "N/A"])
    ws.append([])
    ws.append(["Phone Number",phone_number if phone_number else "N/A"])
    ws.append([])
    ws.append(["Audio URL", audio_url if audio_url else "N/A"])

    # 🔹 Save Excel to Memory
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    # 🔹 Use StreamingResponse to return the in-memory file
    return StreamingResponse(
        excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={call_id}.xlsx"}
    )
